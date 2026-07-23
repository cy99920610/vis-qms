"""Excel/PDF export helpers for the Document Control / Maintenance Tool.
Read-only: builds a workbook/PDF in memory from a queryset or a findings
list — never touches a Document row or an uploaded file."""
import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Table, TableStyle

NAVY = "101B3F"
GOLD = "9C7A29"

REGISTER_COLUMNS = ["Code", "Title", "Revision", "Section", "Folder", "Issue Date", "Format", "Uploaded By"]
FINDINGS_COLUMNS = ["Severity", "Category", "Code", "Title", "Folder", "Message"]


def _register_rows(queryset):
    rows = []
    for d in queryset:
        ext = d.file.name.rsplit(".", 1)[-1].upper() if "." in d.file.name else ""
        uploader = d.uploaded_by.get_full_name() or d.uploaded_by.username if d.uploaded_by else ""
        rows.append([
            d.code, d.title, d.revision, d.get_section_display(),
            d.folder, d.issue_date.strftime("%d/%m/%Y") if d.issue_date else "", ext, uploader,
        ])
    return rows


def _findings_rows(findings):
    return [
        [f["severity"].upper(), f["category_label"], f["code"], f["title"], f["folder"], f["message"]]
        for f in findings
    ]


def _write_xlsx(title, columns, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.sheet_view.showGridLines = False

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    header_cell = ws.cell(row=1, column=1, value=f"VIS-RECRUIT — {title}")
    header_cell.font = Font(bold=True, size=13, color="FFFFFF")
    header_cell.fill = PatternFill("solid", fgColor="808080")
    header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    header_row = 3
    for i, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=i, value=col)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    r = header_row + 1
    for row in rows:
        for i, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.fill = PatternFill("solid", fgColor="F5F7FC" if (r - header_row) % 2 == 0 else "FFFFFF")
        r += 1

    widths = {1: 16, 2: 34, 3: 12, 4: 18, 5: 30, 6: 13, 7: 10, 8: 18}
    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 18)
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _write_pdf(title, columns, rows):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=14 * mm, rightMargin=14 * mm, topMargin=14 * mm, bottomMargin=14 * mm,
    )
    styles = {
        "title": ParagraphStyle("title", fontSize=14, textColor=colors.HexColor("#" + NAVY), spaceAfter=4, fontName="Helvetica-Bold"),
        "sub": ParagraphStyle("sub", fontSize=9, textColor=colors.HexColor("#" + GOLD), spaceAfter=10),
        "cell": ParagraphStyle("cell", fontSize=7.5, leading=10),
        "head": ParagraphStyle("head", fontSize=8, textColor=colors.white, fontName="Helvetica-Bold"),
    }
    elements = [
        Paragraph("VIS-RECRUIT — VIS-QMS Document Control", styles["title"]),
        Paragraph(title, styles["sub"]),
    ]

    table_data = [[Paragraph(c, styles["head"]) for c in columns]]
    for row in rows:
        table_data.append([Paragraph(str(v) if v not in (None, "") else "—", styles["cell"]) for v in row])

    col_count = len(columns)
    page_width = landscape(A4)[0] - 28 * mm
    col_widths = [page_width / col_count] * col_count

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#" + NAVY)),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return buf


def register_xlsx(queryset):
    return _write_xlsx("Final Approved PDF Document Register", REGISTER_COLUMNS, _register_rows(queryset))


def register_pdf(queryset):
    return _write_pdf("Final Approved PDF Document Register", REGISTER_COLUMNS, _register_rows(queryset))


def findings_xlsx(findings):
    return _write_xlsx("Document Control Watchdog — Findings", FINDINGS_COLUMNS, _findings_rows(findings))


def findings_pdf(findings):
    return _write_pdf("Document Control Watchdog — Findings", FINDINGS_COLUMNS, _findings_rows(findings))
